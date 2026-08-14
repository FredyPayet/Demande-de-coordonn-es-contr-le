# -*- coding: utf-8 -*-
"""
Application Streamlit d'automatisation :
  1) Lecture du "tableau 1" (export de contrôle, une ligne = une opération)
  2) Report des colonnes "RAISON SOCIALE du demandeur" -> "SIRET de l'entreprise
     ayant réalisé l'opération" dans les matrices "tableau 2" (une matrice = un
     modèle de fiche CEE, ex. BAR-TH-104, BAR-EN-105, ...)
  3) Génération d'un fichier tableau 2 rempli PAR CLIENT et PAR FICHE, avec
     uniquement les opérations de ce client pour cette fiche
  4) Pour chaque client : boutons pour copier l'objet du mail, copier le
     corps du mail type, et télécharger le tableau 2 correspondant

Lancement :
    pip install streamlit openpyxl pandas
    streamlit run app.py
"""

import io
import os
import re
import base64
import json
import zipfile
import unicodedata
from collections import defaultdict
from datetime import datetime, timedelta

import streamlit as st
import streamlit.components.v1 as components
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


# ---------------------------------------------------------------------------
# Constantes métier
# ---------------------------------------------------------------------------

# Colonnes à NE JAMAIS remplir dans les matrices, même si un nom de colonne
# source correspondait par erreur.
COLONNES_EXCLUES = [
    "numero de telephone du beneficiaire",
    "adresse de courriel du beneficiaire",
    "montant du role actif et incitatif",
    "fonction",
]

# Nom (normalisé) de la colonne source qui contient le code de la fiche
# (ex : BAR-TH-104) et qui sert à choisir la bonne matrice.
COLONNE_CODE_FICHE = "reference de la fiche d'operation standardisee"

# Séparateur utilisé lors de la concaténation de plusieurs colonnes sources
# (ex: "Marque isolant" + "Référence isolant")
SEPARATEUR_CONCATENATION = " "


# Nom (normalisé) de la colonne qui identifie le client (raison sociale du
# bénéficiaire de l'opération). Utilisé pour le regroupement et le mail.
COLONNE_CLIENT = "raison sociale du beneficiaire de l'operation"

FEUILLE_MATRICE = "Personnes morales"  # nom de la feuille dans les matrices

MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MIME_ZIP = "application/zip"


def extraire_numero_lot(nom_fichier_source):
    """Extrait le numéro de lot depuis le nom du fichier tableau 1
    (ex: 'ODICEE_Export_lot_de_controle_XX818P_...xlsx' -> 'XX818P').
    Retourne None si le motif n'est pas trouvé dans le nom du fichier."""
    base = os.path.splitext(os.path.basename(nom_fichier_source))[0]
    m = re.search(r"lot[_\s]*de[_\s]*controle[_\s]*([A-Za-z0-9]+)", base, re.IGNORECASE)
    if m:
        return m.group(1)
    return None


# ---------------------------------------------------------------------------
# Fonctions utilitaires de normalisation / correspondance de colonnes
# ---------------------------------------------------------------------------

def normaliser(texte):
    """Normalise un texte pour comparaison : minuscule, sans accent, sans
    parenthèses, espaces multiples réduits."""
    if texte is None:
        return ""
    s = str(texte)
    s = s.replace("\u2019", "'")
    # retire tout contenu entre parenthèses (ex: "(sous la forme : XXX-XX-XX)")
    s = re.sub(r"\([^)]*\)", " ", s)
    # retire les accents
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower()
    s = re.sub(r"[\s\n\r]+", " ", s).strip()
    return s


def colonne_exclue(nom_normalise):
    return any(nom_normalise.startswith(exc) for exc in COLONNES_EXCLUES)


def trouver_meilleure_correspondance(nom_cible, index_source):
    """Cherche le nom de colonne source correspondant à nom_cible (déjà
    normalisés). Tolère les suffixes/préfixes différents (ex: 'reference
    interne de l'operation' vs '... du demandeur')."""
    if nom_cible in index_source:
        return nom_cible
    # correspondance par préfixe dans les deux sens
    candidats = []
    for cle in index_source:
        if not cle or not nom_cible:
            continue
        if cle.startswith(nom_cible) or nom_cible.startswith(cle):
            candidats.append(cle)
    if len(candidats) == 1:
        return candidats[0]
    if len(candidats) > 1:
        # on prend le plus proche en longueur
        candidats.sort(key=lambda c: abs(len(c) - len(nom_cible)))
        return candidats[0]
    return None


# ---------------------------------------------------------------------------
# Lecture du tableau 1 (source)
# ---------------------------------------------------------------------------

def lire_tableau_source(fichier):
    """Retourne (liste_de_dicts_operations, index_colonnes_normalisees,
    nom_feuille) depuis le fichier Excel source. Détecte automatiquement la
    ligne d'en-têtes (celle où figure 'RAISON SOCIALE du demandeur' en
    colonne A ou plus loin) et la première ligne de données."""
    wb = openpyxl.load_workbook(fichier, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row = None
    for r in range(1, min(10, ws.max_row) + 1):
        for c in range(1, min(80, ws.max_column) + 1):
            v = normaliser(ws.cell(row=r, column=c).value)
            if v == "raison sociale du demandeur":
                header_row = r
                break
        if header_row:
            break
    if header_row is None:
        raise ValueError(
            "Impossible de trouver la ligne d'en-têtes (colonne "
            "'RAISON SOCIALE du demandeur') dans le tableau 1."
        )

    # index colonne_normalisee -> numero de colonne (on garde la 1ere occurrence)
    index_cols = {}
    for c in range(1, ws.max_column + 1):
        v = normaliser(ws.cell(row=header_row, column=c).value)
        if v and v not in index_cols:
            index_cols[v] = c

    operations = []
    for r in range(header_row + 1, ws.max_row + 1):
        # ligne vide -> on l'ignore (mais on continue de scanner jusqu'au bout)
        if all(ws.cell(row=r, column=c).value in (None, "") for c in range(1, ws.max_column + 1)):
            continue
        ligne = {}
        for nom, c in index_cols.items():
            ligne[nom] = ws.cell(row=r, column=c).value
        ligne["_ligne_source"] = r
        operations.append(ligne)

    return operations, index_cols, header_row


# ---------------------------------------------------------------------------
# Détection des codes fiches à partir des noms de fichiers matrices
# ---------------------------------------------------------------------------

_FULL_CODE = re.compile(r"(BAR|BAT)[\s_-](EN|TH)[\s_-](\d{3})(-SE)?", re.IGNORECASE)
_BARE_NUM = re.compile(r"(?<!\d)(\d{3})(?!\d)")


def extraire_codes_fiches(nom_fichier):
    """Devine, à partir du nom du fichier matrice, la ou les codes de fiches
    CEE couverts (ex: 'BAR-TH-145 et 164.xlsx' -> ['BAR-TH-145','BAR-TH-164'])."""
    base = os.path.splitext(os.path.basename(nom_fichier))[0]
    norm = base.replace("_", " ")
    norm = re.sub(r"\bet\b", " ", norm, flags=re.IGNORECASE)
    norm = re.sub(r"\s+", " ", norm)

    tokens = []
    full_spans = []
    for m in _FULL_CODE.finditer(norm):
        tokens.append((m.start(), "full", m))
        full_spans.append((m.start(), m.end()))
    for m in _BARE_NUM.finditer(norm):
        if any(m.start() >= s and m.end() <= e for s, e in full_spans):
            continue
        tokens.append((m.start(), "bare", m))
    tokens.sort(key=lambda t: t[0])

    codes, last_fam = [], None
    for _, typ, m in tokens:
        if typ == "full":
            fam = f"{m.group(1).upper()}-{m.group(2).upper()}"
            se = "-SE" if m.group(4) else ""
            codes.append(f"{fam}-{m.group(3)}{se}")
            last_fam = fam
        elif last_fam:
            codes.append(f"{last_fam}-{m.group(1)}")
    # dédoublonne en conservant l'ordre
    vu, resultat = set(), []
    for c in codes:
        if c not in vu:
            vu.add(c)
            resultat.append(c)
    return resultat


def construire_mapping_fiches(fichiers_matrices):
    """fichiers_matrices : dict {nom_fichier: bytes}. Retourne une liste de
    lignes {code_fiche, fichier_matrice} pour affichage/édition dans l'appli."""
    lignes = []
    for nom in sorted(fichiers_matrices.keys()):
        codes = extraire_codes_fiches(nom)
        if not codes:
            lignes.append({"code_fiche": "(non détecté)", "fichier_matrice": nom})
        for code in codes:
            lignes.append({"code_fiche": code, "fichier_matrice": nom})
    return lignes


# ---------------------------------------------------------------------------
# Lecture des en-têtes d'une matrice (tableau 2) + détection ligne d'en-tête
# ---------------------------------------------------------------------------

def lire_entetes_matrice(contenu_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(contenu_bytes), data_only=False)
    nom_feuille = wb.sheetnames[0]
    for candidat in wb.sheetnames:
        if normaliser(candidat) == normaliser(FEUILLE_MATRICE):
            nom_feuille = candidat
            break
    ws = wb[nom_feuille]

    header_row = None
    for r in range(1, min(10, ws.max_row) + 1):
        v = normaliser(ws.cell(row=r, column=1).value)
        if v == "raison sociale du demandeur":
            header_row = r
            break
    if header_row is None:
        return None, None, None

    entetes = {}
    for c in range(1, ws.max_column + 1):
        v = normaliser(ws.cell(row=header_row, column=c).value)
        if v:
            entetes[c] = v
    return nom_feuille, header_row, entetes


# ---------------------------------------------------------------------------
# Génération du mail type (objet + corps) par client / fiche
# ---------------------------------------------------------------------------

# Nom (normalisé) de la colonne "REFERENCE interne de l'opération" du tableau 1
# (utilisée pour la liste des dossiers concernés, 6 premiers chiffres)
COLONNE_REFERENCE_INTERNE = "reference interne de l'operation du demandeur"

MODELE_CORPS_MAIL = (
    "Bonjour,\n\n"
    "En vue de la réalisation du contrôle sur site à la suite de la fiche "
    "« {fiche_cee} » merci de nous fournir les coordonnées d'un contact afin "
    "de les transmettre à l'organisme de contrôle « {organisme_controle} » "
    "pour la prise de rendez-vous.\n\n"
    "Pour cela merci de remplir les colonnes {colonnes_interlocuteur} du "
    "tableau 2 dans le tableau ci-joint.\n\n"
    "- \"L'interlocuteur - contrôle sur site\" (colonnes bleues) est le contact "
    "qui sera appelé par le bureau de contrôle pour la prise et la réalisation "
    "des rendez-vous. (Obligatoire) Il doit s'agir de la personne qui pourra "
    "faire accéder aux travaux :\n"
    "        • Locataire si les travaux sont accessibles depuis les parties "
    "privatives\n"
    "        • Personne de votre organisme si les travaux sont accessibles "
    "depuis les parties communes\n\n"
    "- \"L'accompagnant - contrôle sur site\" (colonnes bleu clair) est un "
    "contact de votre organisme souhaitant être présent lors du rendez-vous "
    "s'il n'est pas interlocuteur - contrôle sur site. (Facultatif)\n\n"
    "De plus cette opération est aussi susceptible de faire l'objet d'un "
    "contrôle par contact (mail ou appel téléphonique).\n\n"
    "- \"L'interlocuteur - contrôle par contact\" (colonnes orange) est la "
    "personne de votre organisme qui sera contacté pour répondre aux "
    "questions sur la réalisation des travaux. (Obligatoire)\n\n"
    "Il vous appartient d'informer le ou les contacts fournis d'un potentiel "
    "contrôle, de la nécessité de le faciliter et de le permettre sans quoi "
    "les CEE ne pourront être délivrés. Le bureau de contrôle est susceptible "
    "de demander des informations sur la réalisation des travaux pour "
    "préparer la visite du contrôleur, le contact transmis devra connaitre "
    "les travaux ou pouvoir se renseigner.\n\n"
    "Un retour est attendu au plus tard le {date_limite}, le contact transmis "
    "sera joint par le bureau de contrôle prochainement, à ce moment la date "
    "du rendez-vous sera convenue. En fonction de l'avancée des contrôles et "
    "de l'échantillonnage, il est toutefois possible que vous ne soyez pas "
    "contacté.\n\n"
    "Voici vos dossiers concernés :\n"
    "{liste_dossiers}\n\n"
    "L'organisme de contrôle va vous demander les pièces justificatives (OS, "
    "devis, facture...), néanmoins cette demande ne vous concerne pas. Nous "
    "avons déjà les documents nécessaires en notre possession et leur "
    "transmettrons au besoin. Vous pouvez répondre que vous les avez mais ne "
    "pourrez pas les fournir au moment du contrôle.\n\n"
    "Vous souhaitant bonne réception et restant à votre disposition.\n\n"
    "Cordialement,"
)


def construire_objet_mail(numero_lot, identifiant_client):
    return f"Demande de coordonnées - {numero_lot} - {identifiant_client}"


# Colonnes source (tableau 1) à essayer, par ordre de préférence, pour
# retrouver le "bureau / organisme de contrôle" d'une opération (certaines
# fiches n'exposent que le SIREN, pas la raison sociale)
NOMS_ORGANISME_CONTROLE_SOURCE = [
    "raison sociale de l'organisme de controle sur site",
    "raison sociale de l'organisme de controle",
    "siren de l'organisme de controle sur site",
    "siren de l'organisme de controle",
]


def extraire_organisme_controle(operations_client, index_cols_source):
    """Retourne le(s) nom(s) d'organisme de contrôle trouvé(s) parmi les
    opérations du client (colonne du tableau 1), joints par ' / ' s'il y en
    a plusieurs. Retourne None si rien n'est trouvé."""
    nom_colonne = None
    for candidat in NOMS_ORGANISME_CONTROLE_SOURCE:
        if candidat in index_cols_source:
            nom_colonne = candidat
            break
    if not nom_colonne:
        return None
    valeurs = []
    for op in operations_client:
        v = op.get(nom_colonne)
        if v not in (None, "") and str(v) not in valeurs:
            valeurs.append(str(v))
    return " / ".join(valeurs) if valeurs else None


def trouver_plage_colonnes_interlocuteur(entetes):
    """Retourne (lettre_debut, lettre_fin) de la plage de colonnes
    'interlocuteur' à remplir par le client, entre 'Nom interlocuteur -
    contrôle sur site' (inclus) et juste avant 'INFORMATIONS
    COMPLEMENTAIRES' (exclu). Retourne (None, None) si introuvable."""
    col_debut = None
    col_fin_infos = None
    for c, nom_entete in entetes.items():
        if nom_entete == "nom interlocuteur - controle sur site" and col_debut is None:
            col_debut = c
        if nom_entete == "informations complementaires":
            col_fin_infos = c
    if col_debut is None or col_fin_infos is None:
        return None, None
    return get_column_letter(col_debut), get_column_letter(col_fin_infos - 1)


def construire_corps_mail(code_fiche, organisme_controle, colonnes_interlocuteur, operations_client):
    if colonnes_interlocuteur == (None, None):
        colonnes_txt = "[à vérifier manuellement : plage de colonnes non détectée automatiquement]"
    else:
        colonnes_txt = f"{colonnes_interlocuteur[0]} à {colonnes_interlocuteur[1]}"

    date_limite = (datetime.now() + timedelta(days=7)).strftime("%d/%m/%Y")

    prefixes = []
    for op in operations_client:
        ref = op.get(COLONNE_REFERENCE_INTERNE)
        if ref:
            prefixe = str(ref).strip()[:6]
            if prefixe and prefixe not in prefixes:
                prefixes.append(prefixe)
    liste_dossiers = "\n".join(f"- {p}" for p in prefixes) if prefixes else "- (aucune référence trouvée)"

    return MODELE_CORPS_MAIL.format(
        fiche_cee=code_fiche,
        organisme_controle=organisme_controle or "[organisme de contrôle non trouvé]",
        colonnes_interlocuteur=colonnes_txt,
        date_limite=date_limite,
        liste_dossiers=liste_dossiers,
    )


# ---------------------------------------------------------------------------
# Correspondance "données techniques" par fiche CEE (dépend de la fiche CEE,
# ex: BAR-EN-103 -> colonnes isolation à remplir depuis le tableau 1).
#
# Intégrée en dur ici (plus besoin de fournir un fichier à chaque lancement).
# Pour ajouter/modifier une règle : ajoutez une entrée à cette liste avec
#   - code_fiche : le code de la fiche CEE (ex "BAR-EN-103")
#   - colonne_cible : le nom (approximatif, tel qu'affiché) de la colonne à
#     remplir dans le tableau 2
#   - expression_source : le(s) nom(s) de colonne(s) du tableau 1 à
#     concaténer avec un "+", ou une valeur fixe si aucune colonne ne
#     correspond (ex: "3.85")
# ---------------------------------------------------------------------------

DONNEES_TECHNIQUES_PAR_FICHE = [
    {"code_fiche": "BAR-EN-103", "colonne_cible": "Surface déclarée dans l'AH/facture (m2)", "expression_source": "Surface isolant (m²)"},
    {"code_fiche": "BAR-EN-103", "colonne_cible": "Marque et référence de l'isolant déclarées", "expression_source": "Marque isolant + Référence isolant"},
    {"code_fiche": "BAR-EN-103", "colonne_cible": "Valeur R ou Lambda déclaré", "expression_source": "Résistance"},
    {"code_fiche": "BAR-EN-103", "colonne_cible": "Epaisseur minimum théorique (mm) (après tassement si isolant soufflé)", "expression_source": "Epaisseur isolant (mm)"},

    {"code_fiche": "BAR-EN-101", "colonne_cible": "Surface déclarée dans l'AH/facture (m2)", "expression_source": "Surface isolant (m²)"},
    {"code_fiche": "BAR-EN-101", "colonne_cible": "Type d'isolant (soufflé/posé/autre) déclaré", "expression_source": "Type de comble"},
    {"code_fiche": "BAR-EN-101", "colonne_cible": "Marque et référence de l'isolant déclarées", "expression_source": "Marque isolant + Référence isolant"},
    {"code_fiche": "BAR-EN-101", "colonne_cible": "Valeur R ou Lambda déclaré", "expression_source": "Résistance"},
    {"code_fiche": "BAR-EN-101", "colonne_cible": "Epaisseur minimum théorique (mm) (après tassement si isolant soufflé)", "expression_source": "Epaisseur isolant (mm)"},

    {"code_fiche": "BAR-EN-102", "colonne_cible": "Surface déclarée dans l'AH/facture (m2)", "expression_source": "Surface isolant (m²)"},
    {"code_fiche": "BAR-EN-102", "colonne_cible": "Marque et référence de l'isolant déclarées", "expression_source": "Marque isolant + Référence isolant"},
    {"code_fiche": "BAR-EN-102", "colonne_cible": "Valeur de résistance thermique  déclarée", "expression_source": "3.85"},

    {"code_fiche": "BAR-EN-105", "colonne_cible": "Surface déclarée dans l'AH/facture (m2)", "expression_source": "Surface isolant (m²)"},
]


def resoudre_expression_technique(expression, index_cols_source):
    """Détermine si 'expression' (ex: 'Marque isolant + Référence isolant' ou
    'Surface isolant (m²)' ou '3.85') correspond à une ou plusieurs colonnes
    du tableau 1, ou s'il s'agit d'une valeur constante à écrire telle quelle.
    Retourne ('colonnes', [noms_colonnes_source]) ou ('constante', valeur)."""
    parties = [p.strip() for p in expression.split("+") if p.strip()]
    if not parties:
        return ("constante", expression)

    colonnes_resolues = []
    for partie in parties:
        cible = trouver_meilleure_correspondance(normaliser(partie), index_cols_source)
        if not cible:
            # au moins une partie ne correspond à aucune colonne source
            # -> on considère l'expression entière comme une valeur fixe
            return ("constante", expression)
        colonnes_resolues.append(cible)
    return ("colonnes", colonnes_resolues)


def valeur_constante_typee(texte):
    """Convertit '3.85' ou '3,85' en float si possible, sinon renvoie le texte."""
    try:
        return float(texte.replace(",", "."))
    except (ValueError, AttributeError):
        return texte


def construire_regles_techniques(lignes_techniques, index_cols_source):
    """Regroupe les règles techniques par code fiche, avec leur résolution
    (colonnes sources à concaténer, ou valeur constante), prêtes à l'emploi."""
    regles_par_fiche = defaultdict(list)
    for ligne in lignes_techniques:
        type_resolution, donnee = resoudre_expression_technique(
            ligne["expression_source"], index_cols_source
        )
        regles_par_fiche[ligne["code_fiche"]].append(
            {
                "colonne_cible_normalisee": normaliser(ligne["colonne_cible"]),
                "colonne_cible_affichage": ligne["colonne_cible"],
                "type": type_resolution,
                "donnee": donnee,
            }
        )
    return regles_par_fiche


def valeur_regle_technique(operation, regle):
    if regle["type"] == "constante":
        return valeur_constante_typee(regle["donnee"])
    # concaténation d'une ou plusieurs colonnes sources
    parties = []
    for nom_col in regle["donnee"]:
        v = operation.get(nom_col)
        if v not in (None, ""):
            parties.append(str(v))
    if not parties:
        return None
    if len(parties) == 1:
        return parties[0]
    return SEPARATEUR_CONCATENATION.join(parties)


# ---------------------------------------------------------------------------
# Génération des matrices remplies (une par client x fiche)
# ---------------------------------------------------------------------------

def _uniformiser_mise_en_forme(cellule, valeur):
    """Applique une mise en forme uniforme (police noire, non grasse) et
    réinitialise le format numérique, pour que toutes les données remplies
    aient le même rendu quelle que soit la mise en forme héritée du modèle
    (certaines cellules du modèle sont pré-formatées en gras/couleur/date)."""
    police_actuelle = cellule.font
    cellule.font = Font(
        name=police_actuelle.name,
        size=police_actuelle.size,
        bold=False,
        italic=police_actuelle.italic,
        color="FF000000",
    )
    if not isinstance(valeur, datetime):
        cellule.number_format = "General"


# Sections (repérées par leur intitulé de sur-en-tête fusionné, ligne 1) dont
# les colonnes VIDES (non remplies par l'application) sont retirées des
# fichiers générés PAR CLIENT. Les colonnes de ces sections qui ont malgré
# tout été remplies (ex: certaines données techniques logées dans la section
# "bureau de contrôle" par exception, comme la surface d'isolant) sont
# conservées.
SECTIONS_A_RETIRER_POUR_CLIENT = [
    "donnees remplies par le bureau de controle",
    "donnees remplies par l'organisme ayant realise le controle par contact",
    "donnees completees par le demandeur",
]


def _colonnes_vides_a_retirer(col_to_section, colonnes_remplies):
    """Retourne, triée du plus grand au plus petit indice, la liste des
    colonnes à supprimer : celles qui appartiennent (d'après col_to_section,
    capturé avant toute modification) à une section listée dans
    SECTIONS_A_RETIRER_POUR_CLIENT et qui ne contiennent aucune donnée
    remplie par l'application (colonnes_remplies)."""
    a_supprimer = []
    for col, (ligne, texte) in col_to_section.items():
        val = normaliser(texte)
        if any(val.startswith(cible) for cible in SECTIONS_A_RETIRER_POUR_CLIENT):
            if col not in colonnes_remplies:
                a_supprimer.append(col)
    return sorted(set(a_supprimer), reverse=True)


def _capturer_bandeaux_ligne1(ws):
    """Retourne {colonne_originale: (ligne, texte)} pour toutes les
    sur-en-têtes fusionnées des lignes 1-2, avant toute suppression de
    colonne (openpyxl perd ces libellés lors d'une suppression de colonnes
    au sein d'une plage fusionnée : on les capture pour les reconstruire
    nous-mêmes ensuite)."""
    col_to_section = {}
    for mc in ws.merged_cells.ranges:
        if mc.min_row > 2:
            continue
        texte = ws.cell(row=mc.min_row, column=mc.min_col).value
        if texte is None:
            continue
        for c in range(mc.min_col, mc.max_col + 1):
            col_to_section[c] = (mc.min_row, texte)
    return col_to_section


def _reconstruire_bandeaux_ligne1(ws, survivants, col_to_section):
    """Reconstruit les sur-en-têtes fusionnées des lignes 1-2 après
    suppression de colonnes, à partir de la correspondance capturée avant
    suppression et de la liste ordonnée des colonnes originales encore
    présentes. Les plages fusionnées existantes doivent déjà avoir été
    retirées (voir _demerger_lignes) avant l'appel."""
    i = 0
    n = len(survivants)
    while i < n:
        info = col_to_section.get(survivants[i])
        if info is None:
            i += 1
            continue
        j = i
        while j + 1 < n and col_to_section.get(survivants[j + 1]) == info:
            j += 1
        ligne, texte = info
        debut, fin = i + 1, j + 1  # positions actuelles (1-indexées)
        if fin > debut:
            ws.merge_cells(start_row=ligne, start_column=debut, end_row=ligne, end_column=fin)
        ws.cell(row=ligne, column=debut, value=texte)
        i = j + 1


def _demerger_lignes_1_2(ws):
    """Retire toutes les fusions de cellules des lignes 1-2 (à faire AVANT
    toute suppression de colonnes : openpyxl gère mal la suppression de
    colonnes au sein d'une plage fusionnée si on tente de démerger après)."""
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row <= 2:
            ws.unmerge_cells(str(mc))


def remplir_matrice(contenu_bytes, operations_client, index_cols_source, regles_techniques=None, simplifier_pour_client=False):
    """Copie la matrice modèle (contenu_bytes) et y ajoute une ligne par
    opération. 'regles_techniques' (optionnel) est la liste des règles
    supplémentaires à appliquer pour LA fiche de ces opérations (colonnes
    techniques propres à la fiche, ex: isolation). Si 'simplifier_pour_client'
    est vrai, les colonnes des sections destinées au bureau de contrôle sont
    supprimées du fichier (voir SECTIONS_A_RETIRER_POUR_CLIENT).
    Retourne (bytes_du_fichier, colonnes_non_mappees, cibles_techniques_non_trouvees,
    entetes_finales) — entetes_finales reflète la mise en page APRÈS une
    éventuelle suppression de colonnes, pour un calcul correct des lettres de
    colonnes dans le mail type."""
    wb = openpyxl.load_workbook(io.BytesIO(contenu_bytes))
    nom_feuille, header_row, entetes = lire_entetes_matrice(contenu_bytes)
    if nom_feuille is None:
        raise ValueError("Ligne d'en-têtes introuvable dans cette matrice.")
    ws = wb[nom_feuille]

    # Au-delà de la colonne SIRET (+ quelques colonnes voisines : téléphone,
    # email, montant... explicitement exclues), les matrices contiennent la
    # partie "audit terrain" du bureau de contrôle, qui n'a pas d'équivalent
    # dans le tableau 1. On ne signale donc les colonnes non reconnues que
    # jusqu'à cette limite, pour ne pas polluer les avertissements.
    col_siret = None
    for c, nom in entetes.items():
        if nom.startswith("siret de l'entreprise ayant realise l'operation"):
            col_siret = c
            break
    limite_signalement = (col_siret + 8) if col_siret else max(entetes.keys(), default=0)

    # correspondance colonne matrice -> colonne source (nom normalisé)
    # On restreint volontairement cette recherche à la zone "identité" de la
    # matrice (jusqu'à la colonne SIRET + quelques colonnes voisines) : au-delà,
    # les intitulés génériques de la partie audit (ex. "Fonction") pourraient
    # sinon être associés par erreur à une colonne du tableau 1 qui n'a rien à
    # voir (ex. "Fonction interlocuteur projet").
    correspondance = {}  # col_matrice -> nom_colonne_source
    colonnes_non_mappees = []
    for col_matrice, nom_entete in entetes.items():
        if col_matrice > limite_signalement:
            continue
        if colonne_exclue(nom_entete):
            continue
        cible = trouver_meilleure_correspondance(nom_entete, index_cols_source)
        if cible:
            correspondance[col_matrice] = cible
        else:
            colonnes_non_mappees.append(nom_entete)

    # règles techniques spécifiques à la fiche (colonnes situées plus loin
    # dans la matrice, ex: "Surface déclarée dans l'AH/facture (m2)")
    regles_par_colonne = {}  # col_matrice -> regle
    cibles_techniques_non_trouvees = []
    for regle in (regles_techniques or []):
        col_matrice = None
        cible_norm = regle["colonne_cible_normalisee"]
        # correspondance exacte d'abord, puis par préfixe (dans les deux sens)
        for c, nom_entete in entetes.items():
            if nom_entete == cible_norm:
                col_matrice = c
                break
        if col_matrice is None:
            for c, nom_entete in entetes.items():
                if nom_entete and (nom_entete.startswith(cible_norm) or cible_norm.startswith(nom_entete)):
                    col_matrice = c
                    break
        if col_matrice:
            regles_par_colonne[col_matrice] = regle
        else:
            cibles_techniques_non_trouvees.append(regle["colonne_cible_affichage"])

    ligne_ecriture = header_row + 1
    for operation in operations_client:
        for col_matrice, nom_source in correspondance.items():
            valeur = operation.get(nom_source)
            cellule = ws.cell(row=ligne_ecriture, column=col_matrice, value=valeur)
            _uniformiser_mise_en_forme(cellule, valeur)
        for col_matrice, regle in regles_par_colonne.items():
            valeur = valeur_regle_technique(operation, regle)
            cellule = ws.cell(row=ligne_ecriture, column=col_matrice, value=valeur)
            _uniformiser_mise_en_forme(cellule, valeur)
        ligne_ecriture += 1

    if simplifier_pour_client:
        col_to_section = _capturer_bandeaux_ligne1(ws)
        _demerger_lignes_1_2(ws)
        colonnes_remplies = set(correspondance.keys()) | set(regles_par_colonne.keys())
        a_supprimer = _colonnes_vides_a_retirer(col_to_section, colonnes_remplies)
        survivants = [c for c in range(1, ws.max_column + 1) if c not in set(a_supprimer)]
        for c in a_supprimer:
            ws.delete_cols(c, 1)
        _reconstruire_bandeaux_ligne1(ws, survivants, col_to_section)

    # en-têtes finales (après suppression éventuelle), pour le calcul correct
    # des lettres de colonnes utilisées dans le mail type
    entetes_finales = {}
    for c in range(1, ws.max_column + 1):
        v = normaliser(ws.cell(row=header_row, column=c).value)
        if v:
            entetes_finales[c] = v

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read(), colonnes_non_mappees, cibles_techniques_non_trouvees, entetes_finales


# ---------------------------------------------------------------------------
# Bloc d'actions par client (composant HTML/JS) : copier l'objet du mail,
# copier le corps du mail, télécharger le tableau 2 — sans passer par le
# serveur Streamlit pour le téléchargement (évite les soucis liés au blocage
# de requêtes réseau par certains navigateurs/extensions), et avec un retour
# visuel (bouton qui passe en vert) une fois l'action effectuée.
# ---------------------------------------------------------------------------

_STYLE_BOUTON = (
    "padding:0.5rem 1rem;border-radius:0.5rem;border:1px solid #d0d3d9;"
    "background:#ffffff;color:#31333F;font-size:14px;cursor:pointer;"
    "font-family:'Source Sans Pro',sans-serif;transition:background 0.15s;"
)


def rendre_bloc_actions_client(cle, client_label, objet, corps_mail, nom_fichier, octets_fichier):
    b64 = base64.b64encode(octets_fichier).decode("ascii")
    objet_js = json.dumps(objet)
    corps_js = json.dumps(corps_mail)
    nom_fichier_js = json.dumps(nom_fichier)
    label_js = json.dumps(client_label)

    html = f"""
    <div style="font-family:'Source Sans Pro',sans-serif;margin-bottom:14px;">
      <div style="font-size:14px;font-weight:600;margin-bottom:6px;color:#31333F;">
        {client_label}
      </div>
      <div style="display:flex;gap:8px;flex-wrap:wrap;">
        <button id="obj_{cle}" style="{_STYLE_BOUTON}">📋 Copier l'objet</button>
        <button id="mail_{cle}" style="{_STYLE_BOUTON}">📋 Copier le mail</button>
        <button id="dl_{cle}" style="{_STYLE_BOUTON}">⬇️ Télécharger le tableau 2</button>
      </div>
    </div>
    <script>
      (function() {{
        const objetTxt = {objet_js};
        const corpsTxt = {corps_js};
        const nomFichier = {nom_fichier_js};
        const b64Data = "{b64}";

        function copierTexte(txt) {{
          if (navigator.clipboard && navigator.clipboard.writeText) {{
            return navigator.clipboard.writeText(txt).catch(function() {{
              copieFallback(txt);
            }});
          }}
          copieFallback(txt);
          return Promise.resolve();
        }}

        function copieFallback(txt) {{
          const ta = document.createElement('textarea');
          ta.value = txt;
          ta.style.position = 'fixed';
          ta.style.left = '-9999px';
          document.body.appendChild(ta);
          ta.focus();
          ta.select();
          try {{ document.execCommand('copy'); }} catch (e) {{}}
          document.body.removeChild(ta);
        }}

        function marquerFait(bouton, texte) {{
          bouton.style.background = '#28a745';
          bouton.style.color = 'white';
          bouton.style.borderColor = '#28a745';
          bouton.innerText = texte;
        }}

        document.getElementById('obj_{cle}').addEventListener('click', function() {{
          const btn = this;
          copierTexte(objetTxt).then(function() {{
            marquerFait(btn, '✓ Objet copié');
          }});
        }});

        document.getElementById('mail_{cle}').addEventListener('click', function() {{
          const btn = this;
          copierTexte(corpsTxt).then(function() {{
            marquerFait(btn, '✓ Mail copié');
          }});
        }});

        document.getElementById('dl_{cle}').addEventListener('click', function() {{
          const byteChars = atob(b64Data);
          const byteNumbers = new Array(byteChars.length);
          for (let i = 0; i < byteChars.length; i++) {{
            byteNumbers[i] = byteChars.charCodeAt(i);
          }}
          const byteArray = new Uint8Array(byteNumbers);
          const blob = new Blob([byteArray], {{type: "{MIME_XLSX}"}});
          const url = URL.createObjectURL(blob);
          const a = document.createElement('a');
          a.href = url;
          a.download = nomFichier;
          document.body.appendChild(a);
          a.click();
          document.body.removeChild(a);
          setTimeout(function() {{ URL.revokeObjectURL(url); }}, 2000);
          marquerFait(this, '✓ Téléchargé');
        }});
      }})();
    </script>
    """
    components.html(html, height=90)


# ---------------------------------------------------------------------------
# Interface Streamlit
# ---------------------------------------------------------------------------

st.set_page_config(page_title="Automatisation fiches CEE par client", layout="wide")
st.title("Automatisation — Tableau 1 → Fiches par client (Tableau 2)")

st.markdown(
    """
Cette application :
1. lit le **tableau 1** (export de contrôle) ;
2. reporte les colonnes *"RAISON SOCIALE du demandeur"* → *"SIRET de l'entreprise
   ayant réalisé l'opération"* dans la bonne **matrice** (tableau 2) selon le code
   de fiche (ex. BAR-TH-104) ;
3. reporte aussi les **colonnes techniques propres à chaque fiche** (ex.
   surface, épaisseur, marque d'isolant...) selon un tableau de correspondance
   que vous fournissez ;
4. **regroupe les opérations par client** et génère un fichier par client et par
   type de fiche ;
5. prépare un **mail type par client** (Outlook local si disponible).
"""
)

if "resultats" not in st.session_state:
    st.session_state["resultats"] = None

# --- Étape 1 : fichiers -----------------------------------------------------
st.header("1. Fichiers")

col1, col2 = st.columns(2)
with col1:
    fichier_source = st.file_uploader(
        "Tableau 1 — export de contrôle (.xlsx)", type=["xlsx"], key="source"
    )
with col2:
    fichiers_matrices_up = st.file_uploader(
        "Matrices (tableau 2) — fichiers .xlsx ou une archive .zip les contenant",
        type=["xlsx", "zip"],
        accept_multiple_files=True,
        key="matrices",
    )

fichiers_matrices = {}  # nom -> bytes
if fichiers_matrices_up:
    for f in fichiers_matrices_up:
        if f.name.lower().endswith(".zip"):
            with zipfile.ZipFile(f) as z:
                for info in z.infolist():
                    if info.filename.lower().endswith(".xlsx") and not info.is_dir():
                        nom_court = os.path.basename(info.filename)
                        if not nom_court or nom_court.startswith("~$"):
                            continue
                        fichiers_matrices[nom_court] = z.read(info.filename)
        else:
            fichiers_matrices[f.name] = f.read()

if fichiers_matrices:
    st.caption(f"{len(fichiers_matrices)} fichier(s) matrice détecté(s).")

# --- Mapping code fiche -> matrice (détection automatique, non affichée) ---
mapping_valide = None
if fichiers_matrices:
    mapping_valide = construire_mapping_fiches(fichiers_matrices)

# --- Données techniques par fiche (intégrées au code, non affichées) -------
lignes_techniques = DONNEES_TECHNIQUES_PAR_FICHE

# --- Étape 3 : analyse -------------------------------------------------------
if fichier_source and fichiers_matrices and mapping_valide:
    st.header("2. Analyse du tableau 1")

    try:
        operations, index_cols_source, header_row_source = lire_tableau_source(fichier_source)
    except Exception as e:
        st.error(f"Erreur de lecture du tableau 1 : {e}")
        st.stop()

    st.success(f"{len(operations)} opération(s) trouvée(s) dans le tableau 1.")

    numero_lot = extraire_numero_lot(fichier_source.name)
    col_lot1, col_lot2 = st.columns([2, 3])
    with col_lot1:
        numero_lot = st.text_input(
            "Numéro de lot détecté (utilisé pour nommer les fichiers générés)",
            value=numero_lot or "",
        )
    if not numero_lot:
        st.warning(
            "Numéro de lot introuvable dans le nom du fichier tableau 1 "
            "(motif attendu : '...lot_de_controle_NUMERO_...'). "
            "Renseignez-le manuellement ci-dessus avant de générer les fichiers."
        )

    if COLONNE_CLIENT not in index_cols_source:
        st.error(
            "Colonne client introuvable dans le tableau 1 "
            "('RAISON SOCIALE du bénéficiaire de l'opération')."
        )
        st.stop()
    if COLONNE_CODE_FICHE not in index_cols_source:
        st.error(
            "Colonne 'REFERENCE DE LA FICHE d'opération standardisée' "
            "introuvable dans le tableau 1."
        )
        st.stop()

    # code_fiche -> fichier_matrice (à partir de la détection automatique)
    code_vers_fichier = {}
    for ligne in mapping_valide:
        code = (ligne.get("code_fiche") or "").strip().upper()
        fichier = ligne.get("fichier_matrice")
        if code and fichier:
            code_vers_fichier[code] = fichier

    # Détecte les codes fiches présents dans le tableau 1 mais non couverts
    # par la détection automatique (fichier matrice absent, ou nom de fichier
    # ambigu). Dans ce cas, on propose une association manuelle plutôt que
    # d'ignorer silencieusement ces opérations.
    codes_presents = set()
    for op in operations:
        c = op.get(COLONNE_CODE_FICHE)
        if c:
            codes_presents.add(str(c).strip().upper())
    codes_non_couverts = sorted(codes_presents - set(code_vers_fichier.keys()))

    if codes_non_couverts:
        st.warning(
            "Certains codes fiches du tableau 1 n'ont pas été associés "
            "automatiquement à un fichier matrice : " + ", ".join(codes_non_couverts) +
            ". Associez-les ci-dessous pour ne pas perdre ces opérations."
        )
        with st.expander("🔧 Associer manuellement les fiches manquantes", expanded=True):
            options_fichiers = ["-- ignorer ces opérations --"] + sorted(fichiers_matrices.keys())
            for code in codes_non_couverts:
                choix = st.selectbox(
                    f"Fichier matrice pour « {code} »",
                    options=options_fichiers,
                    key=f"mapping_manuel_{code}",
                )
                if choix != options_fichiers[0]:
                    code_vers_fichier[code] = choix

    # règles techniques par fiche (colonnes propres à chaque fiche CEE)
    regles_techniques_par_fiche = (
        construire_regles_techniques(lignes_techniques, index_cols_source)
        if lignes_techniques
        else {}
    )

    # Regroupement client -> fiche -> [operations]
    groupes = defaultdict(lambda: defaultdict(list))
    codes_sans_matrice = set()
    for op in operations:
        client = (op.get(COLONNE_CLIENT) or "(client non renseigné)").strip() \
            if isinstance(op.get(COLONNE_CLIENT), str) else (op.get(COLONNE_CLIENT) or "(client non renseigné)")
        code_fiche_brut = op.get(COLONNE_CODE_FICHE)
        code_fiche = str(code_fiche_brut).strip().upper() if code_fiche_brut else ""
        if code_fiche not in code_vers_fichier:
            codes_sans_matrice.add(code_fiche or "(vide)")
            continue
        groupes[client][code_fiche].append(op)

    if codes_sans_matrice:
        st.info(
            "Opérations ignorées (fiche laissée sans association ci-dessus) : "
            + ", ".join(sorted(codes_sans_matrice))
        )

    # tableau récapitulatif
    recap = []
    for client, par_fiche in groupes.items():
        for code_fiche, ops in par_fiche.items():
            recap.append(
                {
                    "Client": client,
                    "Code fiche": code_fiche,
                    "Fichier matrice": code_vers_fichier[code_fiche],
                    "Nb opérations": len(ops),
                }
            )
    st.dataframe(recap, use_container_width=True)

    # Regroupement par fiche uniquement, tous clients confondus (pour la
    # version "tableau 2 complet")
    groupes_toutes_operations = defaultdict(list)
    for client, par_fiche in groupes.items():
        for code_fiche, ops in par_fiche.items():
            groupes_toutes_operations[code_fiche].extend(ops)

    st.header("3. Génération des fichiers")

    # Style vert pour les boutons de génération une fois l'action effectuée
    # (ciblage CSS via la clé du bouton, propre à Streamlit récent)
    style_vert = ""
    if st.session_state.get("genere_par_client"):
        style_vert += """
        .st-key-btn_generer_client button {
            background-color: #28a745 !important;
            color: white !important;
            border-color: #28a745 !important;
        }
        """
    if st.session_state.get("genere_complet"):
        style_vert += """
        .st-key-btn_generer_complet button {
            background-color: #28a745 !important;
            color: white !important;
            border-color: #28a745 !important;
        }
        """
    if style_vert:
        st.markdown(f"<style>{style_vert}</style>", unsafe_allow_html=True)

    col_gen1, col_gen2 = st.columns(2)
    with col_gen1:
        generer_par_client = st.button(
            "✓ Fichiers générés" if st.session_state.get("genere_par_client") else "Générer les fichiers tableau 2 par client",
            type="primary",
            key="btn_generer_client",
        )
    with col_gen2:
        generer_complet = st.button(
            "✓ Tableau complet généré" if st.session_state.get("genere_complet") else "Générer le tableau 2 complet (tous clients)",
            key="btn_generer_complet",
        )

    if generer_par_client:
        st.session_state["genere_par_client"] = True
    if generer_complet:
        st.session_state["genere_complet"] = True

    if generer_par_client:
        resultats = {}  # client -> [ {nom_fichier, octets, code_fiche, objet, corps_mail} ]
        avertissements_colonnes = {}
        avertissements_technique = {}
        avertissements_mail = {}
        lot_safe = re.sub(r"[\\/:*?\"<>|]+", "_", str(numero_lot or "LOT_INCONNU")).strip()
        with st.spinner("Génération en cours..."):
            for client, par_fiche in groupes.items():
                fichiers_client = []
                plusieurs_fiches = len(par_fiche) > 1
                for code_fiche, ops in par_fiche.items():
                    nom_matrice = code_vers_fichier[code_fiche]
                    contenu = fichiers_matrices[nom_matrice]
                    regles_fiche = regles_techniques_par_fiche.get(code_fiche, [])
                    try:
                        octets, colonnes_non_mappees, cibles_non_trouvees, entetes_finales = remplir_matrice(
                            contenu, ops, index_cols_source, regles_fiche, simplifier_pour_client=True
                        )
                    except Exception as e:
                        st.error(f"Erreur sur {client} / {code_fiche} : {e}")
                        continue
                    client_safe = re.sub(r"[\\/:*?\"<>|]+", "_", str(client))[:80]
                    nom_sortie = f"Demande de coordonnées - {lot_safe} - {client_safe}"
                    if plusieurs_fiches:
                        # un client avec plusieurs types de fiches nécessite un
                        # fichier par fiche : on ajoute le code fiche pour éviter
                        # que les fichiers ne s'écrasent entre eux
                        nom_sortie += f" - {code_fiche}"
                    nom_sortie += ".xlsx"

                    # --- éléments du mail type pour ce client / cette fiche ---
                    # (basés sur les en-têtes APRÈS suppression des colonnes,
                    # pour que les lettres de colonnes citées dans le mail
                    # correspondent bien au fichier réellement envoyé)
                    organisme_controle = extraire_organisme_controle(ops, index_cols_source)
                    colonnes_interlocuteur = trouver_plage_colonnes_interlocuteur(entetes_finales)
                    objet = construire_objet_mail(numero_lot or "LOT_INCONNU", client)
                    corps_mail = construire_corps_mail(code_fiche, organisme_controle, colonnes_interlocuteur, ops)

                    cle_avert = f"{client} / {code_fiche}"
                    manques = []
                    if not organisme_controle:
                        manques.append("organisme de contrôle non trouvé dans le tableau 1")
                    if colonnes_interlocuteur == (None, None):
                        manques.append("plage de colonnes interlocuteur non détectée dans la matrice")
                    if manques:
                        avertissements_mail[cle_avert] = manques

                    fichiers_client.append(
                        {
                            "nom_fichier": nom_sortie,
                            "octets": octets,
                            "code_fiche": code_fiche,
                            "objet": objet,
                            "corps_mail": corps_mail,
                        }
                    )
                    if colonnes_non_mappees:
                        avertissements_colonnes[nom_matrice] = colonnes_non_mappees
                    if cibles_non_trouvees:
                        avertissements_technique[f"{code_fiche} ({nom_matrice})"] = cibles_non_trouvees
                resultats[client] = fichiers_client

        st.session_state["resultats"] = resultats
        st.session_state["avertissements_colonnes"] = avertissements_colonnes
        st.session_state["avertissements_technique"] = avertissements_technique
        st.session_state["avertissements_mail"] = avertissements_mail
        st.rerun()

    if generer_complet:
        resultats_complet = {}  # code_fiche -> (nom_fichier, bytes)
        avertissements_colonnes_c = {}
        avertissements_technique_c = {}
        lot_safe = re.sub(r"[\\/:*?\"<>|]+", "_", str(numero_lot or "LOT_INCONNU")).strip()
        plusieurs_fiches_global = len(groupes_toutes_operations) > 1
        with st.spinner("Génération du tableau 2 complet en cours..."):
            for code_fiche, ops in groupes_toutes_operations.items():
                nom_matrice = code_vers_fichier[code_fiche]
                contenu = fichiers_matrices[nom_matrice]
                regles_fiche = regles_techniques_par_fiche.get(code_fiche, [])
                try:
                    octets, colonnes_non_mappees, cibles_non_trouvees, _ = remplir_matrice(
                        contenu, ops, index_cols_source, regles_fiche
                    )
                except Exception as e:
                    st.error(f"Erreur sur {code_fiche} : {e}")
                    continue
                nom_sortie = lot_safe
                if plusieurs_fiches_global:
                    # plusieurs types de fiches -> un fichier par fiche, on
                    # ajoute le code fiche pour éviter que les fichiers ne
                    # s'écrasent entre eux
                    nom_sortie += f" - {code_fiche}"
                nom_sortie += ".xlsx"
                resultats_complet[code_fiche] = (nom_sortie, octets)
                if colonnes_non_mappees:
                    avertissements_colonnes_c[nom_matrice] = colonnes_non_mappees
                if cibles_non_trouvees:
                    avertissements_technique_c[f"{code_fiche} ({nom_matrice})"] = cibles_non_trouvees

        st.session_state["resultats_complet"] = resultats_complet
        st.session_state["avertissements_colonnes_c"] = avertissements_colonnes_c
        st.session_state["avertissements_technique_c"] = avertissements_technique_c
        st.rerun()

    # --- Retour persistant après génération (survit au rerun immédiat) -----
    if st.session_state.get("genere_par_client"):
        st.success("Génération par client terminée.")
        for nom_matrice, cols in st.session_state.get("avertissements_colonnes", {}).items():
            with st.expander(f"⚠️ Colonnes non reconnues — {nom_matrice}"):
                st.write(", ".join(cols))
        for cle, cibles in st.session_state.get("avertissements_technique", {}).items():
            with st.expander(f"⚠️ Colonnes techniques non trouvées — {cle}"):
                st.write(", ".join(cibles))
        for cle, manques in st.session_state.get("avertissements_mail", {}).items():
            with st.expander(f"⚠️ Informations manquantes pour le mail — {cle}"):
                st.write(", ".join(manques))

    if st.session_state.get("genere_complet"):
        st.success("Génération du tableau 2 complet terminée.")
        for nom_matrice, cols in st.session_state.get("avertissements_colonnes_c", {}).items():
            with st.expander(f"⚠️ Colonnes non reconnues — {nom_matrice}"):
                st.write(", ".join(cols))
        for cle, cibles in st.session_state.get("avertissements_technique_c", {}).items():
            with st.expander(f"⚠️ Colonnes techniques non trouvées — {cle}"):
                st.write(", ".join(cibles))

# --- Étape 5 : téléchargement + mails ---------------------------------------
resultats = st.session_state.get("resultats")
resultats_complet = st.session_state.get("resultats_complet")

if resultats or resultats_complet:
    st.header("4. Téléchargement")

if resultats_complet:
    st.subheader("4a. Tableau 2 complet (toutes les opérations, tous clients)")
    st.caption(
        "Un fichier par fiche, contenant toutes les opérations de tous les "
        "clients — utile pour une vue d'ensemble ou un contrôle global."
    )
    if len(resultats_complet) == 1:
        (nom_fichier, octets) = next(iter(resultats_complet.values()))
        st.download_button(
            f"📄 Télécharger : {nom_fichier}",
            data=octets,
            file_name=nom_fichier,
            mime=MIME_XLSX,
            key="dl_complet_unique",
        )
    else:
        zip_complet = io.BytesIO()
        with zipfile.ZipFile(zip_complet, "w", zipfile.ZIP_DEFLATED) as z:
            for code_fiche, (nom_fichier, octets) in resultats_complet.items():
                z.writestr(nom_fichier, octets)
        zip_complet.seek(0)
        st.download_button(
            "📦 Télécharger le tableau 2 complet (.zip, toutes fiches)",
            data=zip_complet,
            file_name=f"tableau2_complet_{datetime.now().strftime('%Y%m%d_%H%M')}.zip",
            mime=MIME_ZIP,
            key="dl_complet_zip",
        )
        with st.expander("Voir / télécharger fichier par fichier (tableau complet)"):
            for code_fiche, (nom_fichier, octets) in resultats_complet.items():
                st.download_button(
                    f"Télécharger : {nom_fichier}",
                    data=octets,
                    file_name=nom_fichier,
                    mime=MIME_XLSX,
                    key=f"dl_complet_{code_fiche}",
                )

if resultats:
    st.subheader("4b. Actions par client")
    st.caption(
        "Pour chaque client (et chaque fiche s'il y en a plusieurs) : copiez "
        "l'objet du mail, copiez le corps du mail type, et téléchargez le "
        "tableau 2 correspondant. Les boutons passent en vert une fois "
        "l'action effectuée."
    )

    for i, (client, fichiers) in enumerate(resultats.items()):
        if not fichiers:
            continue
        for j, entree in enumerate(fichiers):
            libelle = client
            if len(fichiers) > 1:
                libelle += f" — {entree['code_fiche']}"
            rendre_bloc_actions_client(
                cle=f"{i}_{j}",
                client_label=libelle,
                objet=entree["objet"],
                corps_mail=entree["corps_mail"],
                nom_fichier=entree["nom_fichier"],
                octets_fichier=entree["octets"],
            )

    with st.expander("Voir le détail des mails générés (texte brut)"):
        for i, (client, fichiers) in enumerate(resultats.items()):
            for entree in fichiers:
                st.markdown(f"**{client} — {entree['code_fiche']}**")
                st.text(f"Objet : {entree['objet']}\n\n{entree['corps_mail']}")
                st.divider()

st.divider()
st.caption(
    "Astuce : le mapping code fiche → fichier matrice, les colonnes non "
    "reconnues et les informations manquantes pour le mail sont affichés "
    "pour vérification — pensez à les contrôler avant un envoi réel aux "
    "clients."
)
